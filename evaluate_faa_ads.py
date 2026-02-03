"""
FAA AD Evaluation Script for PC-12 Fleet
=========================================
Evaluates FAA ADs from JSON reports against fleet and writes to Excel.
Creates FAA_AD_Register sheet in AD_Evaluation_Master_CAMO.xlsx

Usage:
    python evaluate_faa_ads.py                    # Evaluate latest report
    python evaluate_faa_ads.py [json_file_path]   # Evaluate specific report
"""

import pandas as pd
from datetime import datetime
import json
import os
import glob
from openpyxl import load_workbook

# Configuration
SKILL_DIR = "C:/Users/delye/.claude/skills/aviation-engineer-agent"
FLEET_FILE = f"{SKILL_DIR}/Fleet_Database_Expanded.xlsx"
MASTER_FILE = f"{SKILL_DIR}/AD_Evaluation_Master_CAMO.xlsx"
FAA_REPORTS_DIR = f"{SKILL_DIR}/FAA_Reports"


def load_fleet():
    """Load fleet database"""
    return pd.read_excel(FLEET_FILE)


def load_faa_report(json_path=None):
    """Load FAA AD report from JSON. If no path given, load latest."""
    if json_path is None:
        # Find latest FAA report
        reports = glob.glob(f"{FAA_REPORTS_DIR}/faa_ads_*.json")
        if not reports:
            print("[ERROR] No FAA reports found in FAA_Reports/")
            return None
        json_path = max(reports, key=os.path.getmtime)
        print(f"[INFO] Loading latest report: {os.path.basename(json_path)}")

    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_models_from_abstract(abstract):
    """Extract PC-12 model variants from abstract text"""
    models = []
    abstract_upper = abstract.upper()

    if 'PC-12/47E' in abstract_upper:
        models.append('PC-12/47E')
    elif 'PC-12/47' in abstract_upper:
        models.append('PC-12/47')
    elif 'PC-12/45' in abstract_upper:
        models.append('PC-12/45')
    elif 'PC-12' in abstract_upper or 'PC12' in abstract_upper:
        # Generic PC-12 - could be all variants
        models = ['PC-12', 'PC-12/45', 'PC-12/47', 'PC-12/47E']

    # Check for PC-24 (different aircraft - should be excluded)
    if 'PC-24' in abstract_upper and not any('PC-12' in m for m in models):
        models = ['PC-24']  # Not our fleet

    return models if models else ['UNKNOWN']


def evaluate_faa_ad_against_fleet(ad_data, fleet, monitoring_date):
    """
    Evaluate a single FAA AD against the fleet.
    Returns list of evaluation rows for Excel.
    """
    results = []
    evaluation_date = datetime.now().strftime('%Y-%m-%d')
    evaluation_time = datetime.now().strftime('%H:%M:%S')

    # Parse AD data
    doc_number = ad_data.get('document_number', '')
    title = ad_data.get('title', '')
    abstract = ad_data.get('abstract', '')
    pub_date = ad_data.get('publication_date', '')
    pdf_url = ad_data.get('pdf_url', '')
    html_url = ad_data.get('html_url', '')
    is_emergency = ad_data.get('is_emergency', False)
    emergency_keyword = ad_data.get('emergency_keyword', '')
    compliance_category = 'EMERGENCY' if is_emergency else 'STANDARD'

    # Extract applicable models from abstract
    applicable_models = parse_models_from_abstract(abstract)

    # Skip if PC-24 only (not our fleet)
    if applicable_models == ['PC-24']:
        print(f"   [SKIP] {doc_number}: PC-24 only (not in fleet)")
        return []

    # Determine ATA chapter from title/abstract (simplified)
    ata_chapter = 'TBD'
    if 'EMERGENCY EXIT' in abstract.upper() or 'PSU' in abstract.upper():
        ata_chapter = 'ATA 25'
    elif 'ENGINE' in abstract.upper() or 'BATTERY' in abstract.upper():
        ata_chapter = 'ATA 24/70'
    elif 'LANDING GEAR' in abstract.upper():
        ata_chapter = 'ATA 32'
    elif 'AIRWORTHINESS LIMITATION' in abstract.upper() or 'ALS' in abstract.upper():
        ata_chapter = 'ATA 05'

    # Extract subject from title
    subject = title.replace('Airworthiness Directives; ', '').strip()
    if len(subject) > 100:
        subject = subject[:97] + '...'

    for _, ac in fleet.iterrows():
        row = {
            # FAA AD Identification
            'AD_Number': f"FAA AD {doc_number}",
            'FR_Document_Number': doc_number,
            'Issuing_Authority': 'FAA',
            'Publication_Date': pub_date,
            'Effective_Date': 'TBD - Proposed Rule',
            'ATA_Chapter': ata_chapter,
            'Subject': subject,

            # Effectivity
            'Applicable_Models': ', '.join(applicable_models),
            'MSN_Range': 'See AD',
            'Affected_PN': 'See AD',
            'Affected_SN': 'See AD',

            # Compliance Requirements
            'Compliance_Category': compliance_category,
            'Compliance_Type': 'TBD',
            'Initial_Compliance': 'TBD - See AD',
            'Recurring_Interval': 'N/A',
            'Terminating_Action': 'See AD',

            # Aircraft Identification
            'Registration': ac['Registration'],
            'MSN': ac['MSN'],
            'Aircraft_Model': ac['Aircraft_Model'],
            'Operator': ac['Operator'],

            # Applicability Evaluation
            'Monitoring_Date': monitoring_date,
            'Evaluation_Date': evaluation_date,
            'Evaluation_Time': evaluation_time,
            'Evaluated_By': 'FAA AD Evaluator v2.0',
        }

        # Determine applicability
        applicability = 'NOT APPLICABLE'
        reason = ''

        # Check if aircraft model is in applicable models
        ac_model = ac['Aircraft_Model']
        model_match = False

        for app_model in applicable_models:
            if app_model in ac_model or ac_model in app_model:
                model_match = True
                break
            # Also check if generic PC-12 applies to all variants
            if app_model == 'PC-12' and 'PC-12' in ac_model:
                model_match = True
                break

        if model_match:
            applicability = 'APPLICABLE'
            reason = f"Model {ac_model} matches effectivity"
        else:
            reason = f"Model {ac_model} not in effectivity ({', '.join(applicable_models)})"

        row['Applicability_Status'] = applicability
        row['Applicability_Reason'] = reason

        # Compliance Status
        if applicability == 'APPLICABLE':
            row['Compliance_Status'] = 'OPEN'
            row['Priority'] = 'EMERGENCY' if is_emergency else 'HIGH'
        else:
            row['Compliance_Status'] = 'N/A'
            row['Priority'] = 'N/A'

        # Compliance tracking (to be filled)
        row['Compliance_Date'] = None
        row['Compliance_Method'] = None
        row['Work_Order_Ref'] = None

        # Resources
        row['Parts_Required'] = 'TBD'
        row['Est_Manhours'] = 'TBD'

        # Links
        row['PDF_URL'] = pdf_url
        row['HTML_URL'] = html_url

        # Audit trail
        row['Check_Date'] = monitoring_date
        row['Evaluator'] = 'FAA AD Evaluator v2.0'

        # Notes
        notes = []
        if is_emergency:
            notes.append(f"EMERGENCY AD - Keyword: {emergency_keyword}")
        if 'PROPOSED' in ad_data.get('type', '').upper():
            notes.append("PROPOSED RULE - Not yet effective")
        row['Notes'] = ' | '.join(notes) if notes else None

        results.append(row)

    return results


def write_to_excel(results_df, master_file=MASTER_FILE):
    """Write FAA evaluation results to Excel FAA_AD_Register sheet"""

    try:
        # Load existing workbook
        book = load_workbook(master_file)

        # Check if FAA_AD_Register sheet exists
        if 'FAA_AD_Register' in book.sheetnames:
            # Load existing data
            existing = pd.read_excel(master_file, sheet_name='FAA_AD_Register')
            # Append new results
            combined = pd.concat([existing, results_df], ignore_index=True)
            # Remove duplicates (same AD + Registration)
            combined = combined.drop_duplicates(
                subset=['AD_Number', 'Registration'],
                keep='last'
            )
        else:
            combined = results_df

        # Write to Excel with both sheets
        with pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            combined.to_excel(writer, sheet_name='FAA_AD_Register', index=False)

        print(f"\n[OK] Results written to {master_file}")
        print(f"     Sheet: FAA_AD_Register")
        print(f"     Total rows: {len(combined)}")

        return True

    except Exception as e:
        print(f"[ERROR] Failed to write Excel: {e}")
        return False


def run_faa_evaluation(json_path=None):
    """Main function to evaluate FAA ADs and write to Excel"""

    print("=" * 70)
    print("FAA AD EVALUATION FOR PC-12 FLEET")
    print("=" * 70)

    # Load fleet
    fleet = load_fleet()
    print(f"[INFO] Fleet loaded: {len(fleet)} aircraft")

    # Load FAA report
    report = load_faa_report(json_path)
    if report is None:
        return None

    monitoring_date = report.get('monitoring_date', datetime.now().strftime('%Y-%m-%d'))
    applicable_ads = report.get('applicable_ads', [])

    print(f"[INFO] Monitoring date: {monitoring_date}")
    print(f"[INFO] Applicable ADs to evaluate: {len(applicable_ads)}")

    if not applicable_ads:
        print("\n[OK] No applicable ADs to evaluate")
        return None

    # Evaluate each AD
    all_results = []

    print("\n" + "-" * 70)
    print("EVALUATING ADs AGAINST FLEET")
    print("-" * 70)

    for ad_data in applicable_ads:
        doc_num = ad_data.get('document_number', 'Unknown')
        is_emergency = ad_data.get('is_emergency', False)
        tag = "[EMERGENCY]" if is_emergency else "[STANDARD]"

        print(f"\n{tag} Evaluating: {doc_num}")
        print(f"   Title: {ad_data.get('title', '')[:60]}...")

        results = evaluate_faa_ad_against_fleet(ad_data, fleet, monitoring_date)

        if results:
            all_results.extend(results)

            # Count applicability
            applicable = sum(1 for r in results if r['Applicability_Status'] == 'APPLICABLE')
            not_applicable = sum(1 for r in results if r['Applicability_Status'] == 'NOT APPLICABLE')

            print(f"   APPLICABLE: {applicable} | NOT APPLICABLE: {not_applicable}")

    if not all_results:
        print("\n[INFO] No fleet-applicable results to write")
        return None

    # Convert to DataFrame
    results_df = pd.DataFrame(all_results)

    # Write to Excel
    print("\n" + "-" * 70)
    print("WRITING TO EXCEL")
    print("-" * 70)

    write_to_excel(results_df)

    # Summary
    print("\n" + "=" * 70)
    print("FAA AD EVALUATION COMPLETE")
    print("=" * 70)

    total_applicable = len(results_df[results_df['Applicability_Status'] == 'APPLICABLE'])
    total_not_applicable = len(results_df[results_df['Applicability_Status'] == 'NOT APPLICABLE'])
    emergency_count = len(results_df[results_df['Compliance_Category'] == 'EMERGENCY'])

    print(f"Total evaluations: {len(results_df)}")
    print(f"APPLICABLE: {total_applicable}")
    print(f"NOT APPLICABLE: {total_not_applicable}")
    if emergency_count > 0:
        print(f"EMERGENCY: {emergency_count} (require same-day action)")

    return results_df


# CLI interface
if __name__ == "__main__":
    import sys

    json_path = None
    if len(sys.argv) > 1:
        json_path = sys.argv[1]

    run_faa_evaluation(json_path)
